from flask import Flask, render_template, request, redirect, url_for, session, jsonify
import json
import os
from datetime import datetime
from werkzeug.utils import secure_filename
import psycopg2
from psycopg2.extras import RealDictCursor
import urllib.parse

app = Flask(__name__)
app.secret_key = "requisiciones2024"

AREAS = [
    {"id": "direccion",      "name": "Dirección",                          "password": "NylexDir3456",   "budget": 74.83,    "cc": 1},
    {"id": "contabilidad",   "name": "Contabilidad y Administración",      "password": "NylexCoad5678",  "budget": 2000.00,  "cc": 2},
    {"id": "licitaciones",   "name": "Licitaciones",                       "password": "NylexLic7896",   "budget": 14412.10, "cc": 4},
    {"id": "obras",          "name": "Obras",                              "password": "NylexOb6432",    "budget": 528.42,   "cc": 5},
    {"id": "almacen",        "name": "Almacén",                            "password": "NylexAlm7653",   "budget": 4853.33,  "cc": 6},
    {"id": "logistica",      "name": "Logística",                          "password": "NylexLog1245",   "budget": 0.00,     "cc": 7},
    {"id": "produccion",     "name": "Producción",                         "password": "NylexProd6912",  "budget": 947.33,   "cc": 8},
    {"id": "diseno",         "name": "Diseño",                             "password": "NylexDis6543",   "budget": 1699.58,  "cc": 9},
    {"id": "compras",        "name": "Compras",                            "password": "NylexComp2918",  "budget": 3488.33,  "cc": 10},
    {"id": "proyectos",      "name": "Proyectos",                          "password": "NylexProy7609",  "budget": 380.00,   "cc": 11},
    {"id": "capital_humano", "name": "Capital Humano",                     "password": "NylexCaph4231",  "budget": 2945.83,  "cc": 12},
    {"id": "seg_higiene",    "name": "Seguridad e Higiene y Mejora Cont.", "password": "NylexSeyh3907",  "budget": 140.00,   "cc": 13},
    {"id": "ti",             "name": "TI",                                 "password": "NylexTI4392",    "budget": 19295.83, "cc": 14},
]

COMPRAS_PASSWORD   = "Compras2024"
DIRECCION_PASSWORD = "Direccion2025"
IMGBB_API_KEY      = "9d97d0d356aaf6b933b1a8ebab8ab14a"

# ── BASE DE DATOS ─────────────────────────────────────────────────────────────
def get_db():
    db_url = os.environ.get("DATABASE_URL")
    print(f"DB URL encontrada: {'Sí' if db_url else 'NO - Variable no existe'}", flush=True)
    if not db_url:
        raise Exception("No se encontró DATABASE_URL")
    if db_url.startswith("postgres://"):
        db_url = db_url.replace("postgres://", "postgresql://", 1)
    conn = psycopg2.connect(db_url, cursor_factory=RealDictCursor)
    return conn

def init_db():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS requisitions (
            id TEXT PRIMARY KEY,
            area TEXT,
            area_id TEXT,
            cc INTEGER,
            productos JSONB,
            notes TEXT,
            total FLOAT,
            authorized BOOLEAN,
            status TEXT,
            created_at TEXT
        );
        CREATE TABLE IF NOT EXISTS catalog (
            id TEXT PRIMARY KEY,
            name TEXT,
            category TEXT,
            unit TEXT,
            img TEXT,
            price FLOAT
        );
        CREATE TABLE IF NOT EXISTS spent (
            area_id TEXT,
            month TEXT,
            amount FLOAT,
            PRIMARY KEY (area_id, month)
        );
    """)
    conn.commit()
    cur.close()
    conn.close()

def load_catalog():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM catalog ORDER BY category, name")
    rows = cur.fetchall()
    cur.close()
    conn.close()
    if not rows:
        return get_default_catalog()
    return [dict(r) for r in rows]

def save_catalog(catalog):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM catalog")
    for p in catalog:
        cur.execute("""
            INSERT INTO catalog (id, name, category, unit, img, price)
            VALUES (%s, %s, %s, %s, %s, %s)
            ON CONFLICT (id) DO UPDATE SET
                name=EXCLUDED.name, category=EXCLUDED.category,
                unit=EXCLUDED.unit, img=EXCLUDED.img, price=EXCLUDED.price
        """, (p["id"], p["name"], p["category"], p["unit"], p["img"], p["price"]))
    conn.commit()
    cur.close()
    conn.close()

def load_requisitions():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM requisitions ORDER BY created_at DESC")
    rows = cur.fetchall()
    cur.close()
    conn.close()
    result = []
    for r in rows:
        d = dict(r)
        d["productos"] = d["productos"] if isinstance(d["productos"], list) else json.loads(d["productos"])
        result.append(d)
    return result

def save_requisition(req):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO requisitions (id, area, area_id, cc, productos, notes, total, authorized, status, created_at)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (req["id"], req["area"], req["area_id"], req["cc"],
          json.dumps(req["productos"], ensure_ascii=False),
          req["notes"], req["total"], req["authorized"],
          req["status"], req["created_at"]))
    conn.commit()
    cur.close()
    conn.close()

def update_status(req_id, status):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE requisitions SET status=%s WHERE id=%s", (status, req_id))
    conn.commit()
    cur.close()
    conn.close()

def get_spent(area_id, month):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT amount FROM spent WHERE area_id=%s AND month=%s", (area_id, month))
    row = cur.fetchone()
    cur.close()
    conn.close()
    return row["amount"] if row else 0

def add_spent(area_id, month, amount):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO spent (area_id, month, amount) VALUES (%s, %s, %s)
        ON CONFLICT (area_id, month) DO UPDATE SET amount = spent.amount + EXCLUDED.amount
    """, (area_id, month, amount))
    conn.commit()
    cur.close()
    conn.close()

def current_month():
    return datetime.now().strftime("%Y-%m")

def get_area(area_id):
    return next((a for a in AREAS if a["id"] == area_id), None)

def get_default_catalog():
    with open("catalogo_backup.json") as f:
        return json.load(f)

# ── INICIALIZAR BD ─────────────────────────────────────────────────────────────
try:
    init_db()
    # Cargar catálogo inicial si está vacío
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) as c FROM catalog")
    count = cur.fetchone()["c"]
    cur.close()
    conn.close()
    if count == 0:
        save_catalog(get_default_catalog())
        print("Catálogo inicial cargado")
except Exception as e:
    print(f"Error BD: {e}")

# ── RUTAS ──────────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return redirect(url_for("login"))

@app.route("/login", methods=["GET","POST"])
def login():
    error = None
    if request.method == "POST":
        mode     = request.form.get("mode")
        password = request.form.get("password")
        if mode == "compras":
            if password == COMPRAS_PASSWORD:
                session["role"] = "compras"
                session["cart"] = {}
                return redirect(url_for("compras"))
            else:
                error = "Contraseña incorrecta."
        else:
            area_id = request.form.get("area_id")
            area    = get_area(area_id)
            if not area:
                error = "Selecciona un área."
            elif password == area["password"]:
                session["role"]    = "area"
                session["area_id"] = area_id
                session["cart"]    = {}
                return redirect(url_for("catalogo"))
            else:
                error = "Contraseña incorrecta."
    return render_template("login.html", areas=AREAS, error=error)

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/catalogo")
def catalogo():
    if session.get("role") != "area":
        return redirect(url_for("login"))
    catalog = load_catalog()
    area    = get_area(session["area_id"])
    mo      = current_month()
    spent   = get_spent(session["area_id"], mo)
    cart    = session.get("cart", {})
    return render_template("catalogo.html", area=area, catalog=catalog, spent=spent, cart=cart)

@app.route("/actualizar_carrito", methods=["POST"])
def actualizar_carrito():
    if session.get("role") != "area":
        return jsonify({"ok": False})
    prod_id = request.form.get("id")
    qty     = int(request.form.get("qty", 0))
    cart    = session.get("cart", {})
    if qty <= 0:
        cart.pop(prod_id, None)
    else:
        cart[prod_id] = qty
    session["cart"] = cart
    session.modified = True
    return jsonify({"ok": True, "count": sum(cart.values())})

@app.route("/carrito")
def carrito():
    if session.get("role") != "area":
        return redirect(url_for("login"))
    catalog = load_catalog()
    area    = get_area(session["area_id"])
    mo      = current_month()
    spent   = get_spent(session["area_id"], mo)
    cart    = session.get("cart", {})
    return render_template("carrito.html", area=area, catalog=catalog, spent=spent, cart=cart)

@app.route("/enviar", methods=["POST"])
def enviar():
    if session.get("role") != "area":
        return redirect(url_for("login"))
    area       = get_area(session["area_id"])
    cart       = session.get("cart", {})
    notes      = request.form.get("notes", "")
    authorized = request.form.get("authorized", "false") == "true"
    dir_pass   = request.form.get("dir_password", "")

    if authorized and dir_pass != DIRECCION_PASSWORD:
        return jsonify({"ok": False, "error": "Contraseña de Dirección incorrecta."})

    catalog     = load_catalog()
    catalog_map = {p["id"]: p for p in catalog}
    productos, total = [], 0
    for pid, qty in cart.items():
        p = catalog_map.get(pid)
        if p and qty > 0:
            subtotal = p["price"] * qty
            total   += subtotal
            productos.append({**p, "qty": qty, "subtotal": subtotal})

    req = {
        "id":         "REQ-" + datetime.now().strftime("%y%m%d%H%M%S"),
        "area":       area["name"],
        "area_id":    area["id"],
        "cc":         area["cc"],
        "productos":  productos,
        "notes":      notes,
        "total":      total,
        "authorized": authorized,
        "status":     "pendiente",
        "created_at": datetime.now().strftime("%d/%m/%Y %H:%M"),
    }
    save_requisition(req)
    add_spent(area["id"], current_month(), total)
    session["cart"] = {}
    session.modified = True
    return jsonify({"ok": True})

@app.route("/compras")
def compras():
    if session.get("role") != "compras":
        return redirect(url_for("login"))
    requisitions = load_requisitions()
    catalog      = load_catalog()
    mo           = current_month()
    spent        = {a["id"]: get_spent(a["id"], mo) for a in AREAS}
    alerts       = [a for a in AREAS if spent[a["id"]] >= a["budget"] and a["budget"] > 0]
    return render_template("compras.html", requisitions=requisitions,
                           areas=AREAS, catalog=catalog,
                           spent=spent, alerts=alerts)

@app.route("/actualizar_status", methods=["POST"])
def actualizar_status():
    if session.get("role") != "compras":
        return jsonify({"ok": False})
    req_id = request.form.get("id")
    status = request.form.get("status")
    update_status(req_id, status)
    return jsonify({"ok": True})

@app.route("/guardar_presupuesto", methods=["POST"])
def guardar_presupuesto():
    if session.get("role") != "compras":
        return jsonify({"ok": False})
    area_id = request.form.get("area_id")
    budget  = float(request.form.get("budget", 0))
    for a in AREAS:
        if a["id"] == area_id:
            a["budget"] = budget
            break
    return jsonify({"ok": True})

@app.route("/guardar_producto", methods=["POST"])
def guardar_producto():
    if session.get("role") != "compras":
        return jsonify({"ok": False})
    prod = {
        "id":       request.form.get("id"),
        "name":     request.form.get("name"),
        "category": request.form.get("category"),
        "unit":     request.form.get("unit"),
        "img":      request.form.get("img", "📦"),
        "price":    float(request.form.get("price", 0)),
    }
    catalog = load_catalog()
    existing = next((i for i,p in enumerate(catalog) if p["id"]==prod["id"]), None)
    if existing is not None:
        catalog[existing] = prod
    else:
        catalog.append(prod)
    save_catalog(catalog)
    return jsonify({"ok": True})

@app.route("/eliminar_producto", methods=["POST"])
def eliminar_producto():
    if session.get("role") != "compras":
        return jsonify({"ok": False})
    prod_id = request.form.get("id")
    catalog = load_catalog()
    catalog = [p for p in catalog if p["id"] != prod_id]
    save_catalog(catalog)
    return jsonify({"ok": True})

@app.route("/subir_imagen", methods=["POST"])
def subir_imagen():
    if session.get("role") != "compras":
        return jsonify({"ok": False, "error": "Sin permisos"})
    if "imagen" not in request.files:
        return jsonify({"ok": False, "error": "No se recibió imagen"})
    file = request.files["imagen"]
    if file.filename == "":
        return jsonify({"ok": False, "error": "No se seleccionó archivo"})
    try:
        import base64, urllib.request as ureq
        image_data = base64.b64encode(file.read()).decode("utf-8")
        data = urllib.parse.urlencode({
            "key": IMGBB_API_KEY,
            "image": image_data
        }).encode()
        req = ureq.Request("https://api.imgbb.com/1/upload", data=data)
        with ureq.urlopen(req, timeout=30) as response:
            result = json.loads(response.read().decode())
        if result.get("success"):
            return jsonify({"ok": True, "url": result["data"]["url"]})
        else:
            return jsonify({"ok": False, "error": str(result)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route("/export_catalog")
def export_catalog():
    if session.get("role") != "compras":
        return redirect(url_for("login"))
    catalog = load_catalog()
    return jsonify(catalog)

@app.route("/export_excel")
def export_excel():
    if session.get("role") != "compras":
        return redirect(url_for("login"))
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except:
        return "pip install openpyxl", 500

    requisitions = load_requisitions()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Requisiciones"

    # Encabezados
    headers = ["ID", "Área", "CC", "Fecha", "Estado", "Total", "Autorizado", "Notas", "Productos"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1e293b")
        cell.alignment = Alignment(horizontal="center")

    for row, r in enumerate(requisitions, 2):
        prods = ", ".join([f"{p['name']} x{p['qty']}" for p in r["productos"]])
        ws.cell(row=row, column=1, value=r["id"])
        ws.cell(row=row, column=2, value=r["area"])
        ws.cell(row=row, column=3, value=r["cc"])
        ws.cell(row=row, column=4, value=r["created_at"])
        ws.cell(row=row, column=5, value=r["status"])
        ws.cell(row=row, column=6, value=r["total"])
        ws.cell(row=row, column=7, value="Sí" if r["authorized"] else "No")
        ws.cell(row=row, column=8, value=r["notes"])
        ws.cell(row=row, column=9, value=prods)

    # Ajustar columnas
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from flask import send_file
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"requisiciones_{datetime.now().strftime('%Y%m%d')}.xlsx")

if __name__ == "__main__":
    app.run(debug=True)